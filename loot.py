import openpyxl

def findClass(classInput):
    if("Priest" in classInput):
        return "priest"
    if("Druid" in classInput):
        return "druid"
    if("Mage" in classInput):
        return "mage"
    if("Rogue" in classInput):
        return "rogue"
    if("Warlock" in classInput):
        return "warlock"
    if("Warrior" in classInput):
        return "warrior"
    if("Shaman" in classInput):
        return "paladin"
    if("Hunter" in classInput):
        return "hunter"

def findTier(itemName, charClass):
    if(charClass == "priest" or "warlock" or "mage"):
        if("Circlet" in itemName):
            return "Desecrated Circlet"
        if("Robe" in itemName):
            return "Desecrated Robe"
        if("Shoulderpads" in itemName):
            return "Desecrated Shoulderpads"
        if("Sandals" in itemName):
            return "Desecrated Sandals"
        if("Bindings" in itemName):
            return "Desecrated Bindings"
        if("Gloves" in itemName):
            return "Desecrated Gloves"
        if("Leggings" in itemName):
            return "Desecrated Leggings"
        if("Belt" in itemName):
            return "Desecrated Belt"
    if(charClass == "paladin" or "hunter" or "druid"):
        if("Headpiece" in itemName):
            return "Desecrated Headpiece"
        if("Tunic" in itemName):
            return "Desecrated Tunic"
        if("Spaulders" in itemName):
            return "Desecrated Spaulders"
        if("Boots" in itemName):
            return "Desecrated Boots"
        if("Wristguards" in itemName):
            return "Desecrated Wristguards"
        if("Handguards" in itemName):
            return "Desecrated Handguards"
        if("Legguards" in itemName):
            return "Desecrated Legguards"
        if("Girdle" in itemName):
            return "Desecrated Girdle"
    if(charClass == "warrior" or "rogue"):
        if("Helmet" in itemName): 
            return "Desecrated Helmet"
        if("Breastplate" in itemName): 
            return "Desecrated Breastplate"
        if("Pauldrons" in itemName): 
            return "Desecrated Pauldrons"
        if("Sabatons" in itemName): 
            return "Desecrated Sabatons"
        if("Wristguards" in itemName): 
            return "Desecrated Wristguards"
        if("Gauntlets" in itemName): 
            return "Desecrated Gauntlets"
        if("Legplates" in itemName): 
            return "Desecrated Legplates"
        if("Waistguard" in itemName): 
            return "Desecrated Waistguard"
    # Find a Ring
    if("Ring" in itemName and "priest" in charClass):
        return "Ring of Faith"
    if("Ring" in itemName and "mage" in charClass):
        return "Frostfire Ring"
    if("Ring" in itemName and "warlock" in charClass):
        return "Plagueheart Ring"
    if("Ring" in itemName and "rogue" in charClass):
        return "Bonescythe Ring"
    if("Ring" in itemName and "druid" in charClass):
        return "Ring of Dreamwalker"
    if("Ring" in itemName and "hunter" in charClass):
        return "Ring of Cryptstalker"
    if("Ring" in itemName and "paladin" in charClass):
        return "Ring of Redemption"
    if("Ring" in itemName and "warrior" in charClass):
        return "Ring of the Dreadnaught"

def otherItemCheck(itemName):
    if("Grasp of the Old God" in itemName):
        return "Husk of the Old God"
    if("Qiraji Bindings of Dominance" in itemName):
        return "Qiraji Bindings of Dominance"
    if("Girdle of Grand Crusader" in itemName):
        return "Belt of the Grand Crusader"
    if("Pauldrons of Grand Crusader" in itemName):
        return "Spaulders of the Grand Crusader"
    if("Creeping Vine Helm" in itemName):
        return "Creeping Vine Helmet"
    else:
        return itemName

# Open character loot sheet and open workbook
pathChar = "C:\Coding\LootAuto\Juggin.xlsx"
wbChar = openpyxl.load_workbook(pathChar)
sheet_char = wbChar.active 

# Get character name
charName = sheet_char.cell(row=2, column = 1).value
# Get character class
charClass = findClass(sheet_char.cell(row=4, column = 1).value)
print("charclass = " + charClass)

# Open master loot sheet
pathMaster = "C:\Coding\LootAuto\LootSheet.xlsx"
wbLootSheet = openpyxl.load_workbook(pathMaster)
wsAtt = wbLootSheet["Attendance"]
charPos = -1
#print(charName)
# Find attendance for specific character
for i in range(1,58):
    if(wsAtt.cell(row=i, column=1).value == charName):
        charPos = i
        break


# Cycle through items on loot sheet
for i in range(5,45):
    prioPos = sheet_char.cell(row=i, column=3).value
    for j in range(4,6):
        # Find items
        #if type(sheet_char.cell(row=i, column=j).value) == str:
        #    print("it's a string! " + sheet_char.cell(row=i, column=j).value)
        if (sheet_char.cell(row=i, column=j).value is not None and type(sheet_char.cell(row=i, column=j).value) == str ):
            itemName = sheet_char.cell(row=i, column=j).value
            #print("THIS IS THE ITEM TO PROCESS: " + itemName)
            # If Qiraji Bindings of Command remove (Shoulder) and (Boots) references
            if("Qiraji Bindings of Command" in itemName):
                itemName = "Qiraji Bindings of Command"
                
            # If item is tier gear
            if("T3" in itemName):
                itemName = findTier(itemName, charClass)
            itemName = otherItemCheck(itemName)
            ##print("searching for {}".format(itemName))
            #print("{} priod at value {}".format(itemName, prioPos))
            # Once item is found, put in sheet
            itemCol = -1
            wsAQ = wbLootSheet["AQ"]
            k = 0
            for k in range(1, 160):
                if(wsAQ.cell(row=1, column=k).value == itemName.strip()):
                    #print("{} found in AQ sheet".format(itemName))
                    sheet = wsAQ
                    itemCol = k
                    break
            wsNaxx = wbLootSheet["Naxx"]
            for k in range(1, 187):
                if(wsNaxx.cell(row=1, column=k).value == itemName.strip()):
                    #print("{} found in Naxx sheet".format(itemName))
                    sheet = wsNaxx
                    itemCol = k
                    break
            
            if(itemCol == -1):
                print("Error finding item: {}".format(itemName))
                break
            else:
                q = 2
                addNewString = "=addNew(\"{}\",{},{})".format(charName, prioPos, charPos)
                inputFlag = False
                while(True):
                    if (sheet.cell(row=q, column=itemCol).value is None):
                        sheet.cell(row=q, column=itemCol).value = addNewString
                        #print("{} is now in the loot sheet".format(itemName))
                        break
                    else:
                        q = q + 1

wbLootSheet.save("C:\Coding\LootAuto\LootSheet.xlsx") 